from datetime import datetime as dt
import os
import openpyxl

def set_column_width(sheet, column):
    max_len = 0
    for cell in actual_sheet[column]:
        if cell.value == None:
            continue
        if len(cell.value) > max_len:
            max_len = len(cell.value)

    actual_sheet.column_dimensions[column].width = max_len

def today_date():
    day_now = dt.now().strftime('%d.%m.')
    month_now = dt.now().strftime('%B')
    year_now = dt.now().strftime('%Y')

    return day_now, month_now, year_now


actions = '''ACTIONS:  [ADD FOOD (add)]  [REMOVE FOOD (rm)]  [B.M. (bm)]  [SHOW MONTH (show)]  [QUIT (q)]:
'''


day_now, month_now, year_now = today_date()

path_to_folder = 'C:\\Users\\viliam\\Desktop\\food_journal\\'
file_name = year_now + '.xlsx'
path = 'C:\\Users\\viliam\\Desktop\\food_journal\\' + year_now + '.xlsx'


if not file_name in os.listdir(path_to_folder): # creates xlsx for actual year, if exists -> opens it
    wb = openpyxl.Workbook()
    wb.save(path)

wb = openpyxl.load_workbook(path)


sheets = wb.sheetnames

if sheets[-1] != month_now: # creates sheet for actual month, if doesn't exist
    ws = wb.create_sheet(month_now)
    wb.save(path)

if 'Sheet' in sheets: # deleting built-in sheet created automaticaly
    del wb['Sheet']

wb.save(path) # creating new files / sheets finished, saving into empty file / sheet

# ---------------------------------------------------------------------------------------------------------------
while True:

    actual_sheet = wb[month_now] # moving to actual month-sheet

    do = input(actions)


    # add food stamp
    if do == 'add':

        data = [day_now]

        time = input('Insert actual time (hh:mm): ')
        food = input('Insert all the food in your meal:  ')

        data += [time, food]

        actual_sheet.append(data)

        #set food column according to maximal column width needed
        set_column_width(actual_sheet, 'C')
        wb.save(path)


    if do == 'rm':
        last = actual_sheet.max_row.__index__()
        actual_sheet.delete_rows(last)
        wb.save(path)


    if do == 'bm':
        last = actual_sheet.max_row.__index__()
        comment = input('Describe how was your job + how many times: ')
        actual_sheet[f'D{last}'] = comment
        wb.save(path)
        set_column_width(actual_sheet, 'D')
        wb.save(path)


    if do == 'show':
        for row in actual_sheet.iter_rows():
            print(f'Day: {row[0].value} | Time: {row[1].value:5} | Food {row[2].value} |')


    if do == 'q':
        quit()

    input('='*20 + '\nPress enter to continue: ')




